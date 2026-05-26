# database.py - Модуль для работы с базой данных

import logging
import sqlite3
import threading
import contextlib
from datetime import datetime, timedelta
from html import escape

import config
from billing import line_totals_with_vat, split_unit_price_with_vat
from constants import OrderStatus, PartStatus, PaymentStatus, VAT_RATE
from repository import OrderRepository

logger = logging.getLogger(__name__)


class ServiceDatabase:
    _connection_counter = 0
    _connection_lock = threading.Lock()

    def __init__(self, db_path=None):
        self.db_path = str(db_path if db_path is not None else config.DB_PATH)
        self.connection = None
        self.orders = OrderRepository(self)
        self._db_lock = threading.RLock()
        self._active_connections = set()
        self._connection_limit = 10
        self.create_database()

    def connect(self):
        """Получить соединение с БД (потокобезопасно)"""
        with self._db_lock:
            self.connection = sqlite3.connect(
                self.db_path,
                timeout=30,
                check_same_thread=False
            )
            self.connection.row_factory = sqlite3.Row
            self.connection.execute('PRAGMA foreign_keys = ON')
            self.connection.execute('PRAGMA journal_mode = WAL')
            conn_id = id(self.connection)
            with ServiceDatabase._connection_lock:
                ServiceDatabase._connection_counter += 1
                self._active_connections.add(conn_id)
        return self.connection

    @contextlib.contextmanager
    def get_connection(self):
        """
        Контекстный менеджер для безопасного получения соединения.
        Гарантирует COMMIT при успехе и ROLLBACK при ошибке.
        Автоматическое закрытие в finally блоке.
        """
        with self._db_lock:
            conn = sqlite3.connect(
                self.db_path,
                timeout=30,
                check_same_thread=False
            )
            conn.row_factory = sqlite3.Row
            conn.execute('PRAGMA foreign_keys = ON')
            conn_id = id(conn)
            with ServiceDatabase._connection_lock:
                ServiceDatabase._connection_counter += 1
                self._active_connections.add(conn_id)
            try:
                yield conn
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass
                raise
            else:
                try:
                    conn.commit()
                except Exception:
                    conn.rollback()
                    raise
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
                with ServiceDatabase._connection_lock:
                    self._active_connections.discard(conn_id)

    def close(self):
        """Закрыть соединение с БД (потокобезопасно)"""
        if self.connection:
            try:
                conn_id = id(self.connection)
                with ServiceDatabase._connection_lock:
                    self._active_connections.discard(conn_id)
                self.connection.close()
            except Exception as e:
                logger.warning("Error closing connection: %s", e)
            finally:
                self.connection = None

    def create_database(self):
        """Create database schema if missing and run migrations."""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('PRAGMA journal_mode = WAL')

            # Таблица: Клиенты
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS clients (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    full_name TEXT NOT NULL,
                    phone TEXT,
                    email TEXT,
                    address TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    notes TEXT,
                    balance REAL DEFAULT 0
                )
            ''')

            # Таблица: Автомобили
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS vehicles (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    client_id INTEGER NOT NULL,
                    vin TEXT UNIQUE NOT NULL,
                    brand TEXT NOT NULL,
                    model TEXT,
                    year INTEGER,
                    license_plate TEXT,
                    color TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (client_id) REFERENCES clients(id)
                )
            ''')

            # Таблица: Заказ-наряды
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS orders (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    order_number INTEGER UNIQUE NOT NULL,
                    vehicle_id INTEGER NOT NULL,
                    client_id INTEGER NOT NULL,
                    complaint TEXT,
                    status TEXT DEFAULT 'Ожидание',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    completed_at TIMESTAMP,
                    total_labor REAL DEFAULT 0,
                    total_parts REAL DEFAULT 0,
                    total_services REAL DEFAULT 0,
                    total_amount REAL DEFAULT 0,
                    notes TEXT,
                    FOREIGN KEY (vehicle_id) REFERENCES vehicles(id),
                    FOREIGN KEY (client_id) REFERENCES clients(id)
                )
            ''')

            # Таблица: Работы (устаревшая, оставлена для совместимости)
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS order_labor (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    order_id INTEGER NOT NULL,
                    description TEXT NOT NULL,
                    price REAL NOT NULL,
                    quantity REAL DEFAULT 1,
                    subtotal REAL NOT NULL,
                    FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE
                )
            ''')

            # Таблица: Запчасти в заказе
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS order_parts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    order_id INTEGER NOT NULL,
                    part_name TEXT NOT NULL,
                    part_number TEXT,
                    quantity REAL NOT NULL,
                    purchase_price REAL DEFAULT 0,
                    selling_price REAL NOT NULL,
                    status TEXT DEFAULT 'Требуется заказ',
                    supplier TEXT,
                    ordered_at TIMESTAMP,
                    received_at TIMESTAMP,
                    notes TEXT,
                    FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE
                )
            ''')

            # Таблица: Услуги/работы (с НДС и без)
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS order_services (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    order_id INTEGER NOT NULL,
                    description TEXT NOT NULL,
                    price_with_vat REAL NOT NULL,
                    vat_amount REAL NOT NULL,
                    price_without_vat REAL NOT NULL,
                    quantity INTEGER DEFAULT 1,
                    subtotal_with_vat REAL NOT NULL,
                    subtotal_without_vat REAL NOT NULL,
                    notes TEXT,
                    FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE
                )
            ''')

            # Таблица: Платежи по заказам
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS order_payments (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    order_id INTEGER NOT NULL,
                    amount REAL NOT NULL,
                    paid_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    note TEXT,
                    applied_amount REAL DEFAULT 0,
                    credit_amount REAL DEFAULT 0,
                    FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE
                )
            ''')

            # Таблица: Поставщики
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS suppliers (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL UNIQUE
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS schema_migrations (
                    version INTEGER PRIMARY KEY,
                    applied_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS order_status_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    order_id INTEGER NOT NULL,
                    old_status TEXT,
                    new_status TEXT NOT NULL,
                    changed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    note TEXT,
                    FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE
                )
            ''')

            # Индексы
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_vehicles_vin ON vehicles(vin)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_vehicles_plate ON vehicles(license_plate)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_clients_phone ON clients(phone)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_orders_status ON orders(status)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_orders_number ON orders(order_number)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_orders_created ON orders(created_at)')

            self.run_migrations(cursor)

        logger.info("Database ready: %s", self.db_path)

    def run_migrations(self, cursor):
        cursor.execute('SELECT version FROM schema_migrations')
        applied = {row[0] for row in cursor.fetchall()}

        if 1 not in applied:
            cursor.execute('UPDATE orders SET status=? WHERE status=?', (OrderStatus.WAITING.value, 'Ожидание'))
            cursor.execute('UPDATE orders SET status=? WHERE status=?', (OrderStatus.IN_PROGRESS.value, 'В работе'))
            cursor.execute('UPDATE orders SET status=? WHERE status=?', (OrderStatus.WAITING_PARTS.value, 'Ждём запчасти'))
            cursor.execute('UPDATE orders SET status=? WHERE status=?', (OrderStatus.DONE.value, 'Готово'))
            cursor.execute('UPDATE order_parts SET status=? WHERE status=?', (PartStatus.NEEDS_ORDER.value, 'Требуется заказ'))
            cursor.execute('INSERT INTO schema_migrations (version) VALUES (1)')

        if 2 not in applied:
            self._add_column_if_missing(cursor, 'orders', 'paid_amount', 'REAL DEFAULT 0')
            self._add_column_if_missing(cursor, 'orders', 'payment_status', f"TEXT DEFAULT '{PaymentStatus.UNPAID.value}'")
            cursor.execute('''
                UPDATE orders
                SET paid_amount = COALESCE(paid_amount, 0),
                    payment_status = CASE
                        WHEN COALESCE(paid_amount, 0) <= 0 THEN ?
                        WHEN COALESCE(paid_amount, 0) >= COALESCE(total_amount, 0) THEN ?
                        ELSE ?
                    END
            ''', (PaymentStatus.UNPAID.value, PaymentStatus.PAID.value, PaymentStatus.PARTIAL.value))
            cursor.execute('INSERT INTO schema_migrations (version) VALUES (2)')

        if 3 not in applied:
            self._add_column_if_missing(cursor, 'clients', 'balance', 'REAL DEFAULT 0')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS order_payments (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    order_id INTEGER NOT NULL,
                    amount REAL NOT NULL,
                    paid_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    note TEXT,
                    applied_amount REAL DEFAULT 0,
                    credit_amount REAL DEFAULT 0,
                    FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE
                )
            ''')
            self._add_column_if_missing(cursor, 'order_payments', 'applied_amount', 'REAL DEFAULT 0')
            self._add_column_if_missing(cursor, 'order_payments', 'credit_amount', 'REAL DEFAULT 0')
            cursor.execute('INSERT INTO schema_migrations (version) VALUES (3)')

    def _add_column_if_missing(self, cursor, table_name, column_name, definition):
        cursor.execute(f'PRAGMA table_info({table_name})')
        columns = {row[1] for row in cursor.fetchall()}
        if column_name not in columns:
            cursor.execute(f'ALTER TABLE {table_name} ADD COLUMN {column_name} {definition}')

    def _ensure_order_payments_columns(self, cursor):
        """Ensure order_payments table has applied_amount and credit_amount columns."""
        self._add_column_if_missing(cursor, 'order_payments', 'applied_amount', 'REAL DEFAULT 0')
        self._add_column_if_missing(cursor, 'order_payments', 'credit_amount', 'REAL DEFAULT 0')

    # ========== КЛИЕНТЫ ==========
    def add_client(self, full_name, phone=None, email=None, address=None, notes=None):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO clients (full_name, phone, email, address, notes)
                VALUES (?, ?, ?, ?, ?)
            ''', (full_name, phone, email, address, notes))
            client_id = cursor.lastrowid
            logger.info("Added client: id=%d, name=%s", client_id, full_name)
            return client_id

    def find_client_by_name(self, full_name):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM clients WHERE full_name = ?', (full_name,))
            client = cursor.fetchone()
            return dict(client) if client else None

    def get_client(self, client_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM clients WHERE id = ?', (client_id,))
            client = cursor.fetchone()
            return dict(client) if client else None

    def update_client(self, client_id, full_name, phone=None, email=None, address=None, notes=None):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE clients
                SET full_name=?, phone=?, email=?, address=?, notes=?
                WHERE id=?
            ''', (full_name, phone, email, address, notes, client_id))
            logger.info("Updated client: id=%d, name=%s", client_id, full_name)

    def get_all_clients(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM clients ORDER BY created_at DESC')
            return [dict(row) for row in cursor.fetchall()]

    # ========== АВТОМОБИЛИ ==========
    def add_vehicle(self, client_id, vin, brand, model=None, year=None,
                    license_plate=None, color=None):
        try:
            with self.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO vehicles (client_id, vin, brand, model, year, license_plate, color)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (client_id, vin, brand, model, year, license_plate, color))
                return cursor.lastrowid
        except sqlite3.IntegrityError:
            return None

    def update_vehicle(self, vehicle_id, client_id, vin, brand, model=None, year=None,
                       license_plate=None, color=None):
        try:
            with self.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    UPDATE vehicles
                    SET client_id=?, vin=?, brand=?, model=?, year=?, license_plate=?, color=?
                    WHERE id=?
                ''', (client_id, vin, brand, model, year, license_plate, color, vehicle_id))
                return True
        except sqlite3.IntegrityError:
            return False

    def find_vehicle_by_vin(self, vin):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT v.*, c.full_name, c.phone
                FROM vehicles v
                JOIN clients c ON v.client_id = c.id
                WHERE v.vin = ?
            ''', (vin,))
            vehicle = cursor.fetchone()
            return dict(vehicle) if vehicle else None

    def get_all_vehicles(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT v.*, c.full_name as owner_name
                FROM vehicles v
                JOIN clients c ON v.client_id = c.id
                ORDER BY v.created_at DESC
            ''')
            return [dict(row) for row in cursor.fetchall()]

    def get_vehicle(self, vehicle_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT v.*, c.full_name as owner_name, c.phone
                FROM vehicles v
                JOIN clients c ON v.client_id = c.id
                WHERE v.id = ?
            ''', (vehicle_id,))
            vehicle = cursor.fetchone()
            return dict(vehicle) if vehicle else None

    # ========== ЗАКАЗЫ ==========
    def create_order(self, vehicle_id, client_id, complaint, order_number=None):
        conn = self.connect()
        cursor = conn.cursor()
        if order_number is None:
            cursor.execute('SELECT COALESCE(MAX(order_number), 0) + 1 FROM orders')
            order_number = cursor.fetchone()[0]
        cursor.execute('''
            INSERT INTO orders (order_number, vehicle_id, client_id, complaint, status)
            VALUES (?, ?, ?, ?, ?)
        ''', (order_number, vehicle_id, client_id, complaint, OrderStatus.WAITING.value))
        order_id = cursor.lastrowid
        cursor.execute('''
            INSERT INTO order_status_history (order_id, old_status, new_status, note)
            VALUES (?, NULL, ?, ?)
        ''', (order_id, OrderStatus.WAITING.value, 'Создание заказа'))
        conn.commit()
        self.close()
        return order_id, order_number

    def get_orders_by_status(self, status=None):
        conn = self.connect()
        cursor = conn.cursor()
        if status:
            cursor.execute('''
                SELECT o.*, v.brand, v.model, v.license_plate, c.full_name
                FROM orders o
                JOIN vehicles v ON o.vehicle_id = v.id
                JOIN clients c ON o.client_id = c.id
                WHERE o.status = ?
                ORDER BY o.created_at DESC
            ''', (status,))
        else:
            cursor.execute('''
                SELECT o.*, v.brand, v.model, v.license_plate, c.full_name
                FROM orders o
                JOIN vehicles v ON o.vehicle_id = v.id
                JOIN clients c ON o.client_id = c.id
                ORDER BY o.created_at DESC
            ''')
        orders = [dict(row) for row in cursor.fetchall()]
        self.close()
        return orders

    def search_orders(self, query: str = '', status=None, vin=None, order_number=None, phone=None):
        """Поиск заказов по имени владельца, регистрационному номеру или VIN"""
        conn = self.connect()
        cursor = conn.cursor()
        params = []
        filters = []
        if query:
            pattern = f"%{query}%"
            filters.append('(c.full_name LIKE ? OR v.license_plate LIKE ? OR v.vin LIKE ? OR c.phone LIKE ? OR CAST(o.order_number AS TEXT) LIKE ?)')
            params.extend([pattern, pattern, pattern, pattern, pattern])
        if status:
            filters.append('o.status = ?')
            params.append(status)
        if vin:
            filters.append('v.vin LIKE ?')
            params.append(f"%{vin}%")
        if order_number:
            filters.append('CAST(o.order_number AS TEXT) LIKE ?')
            params.append(f"%{order_number}%")
        if phone:
            filters.append('c.phone LIKE ?')
            params.append(f"%{phone}%")
        where = f"WHERE {' AND '.join(filters)}" if filters else ''
        cursor.execute(f'''
            SELECT o.*, v.brand, v.model, v.license_plate, v.vin, c.full_name
            FROM orders o
            JOIN vehicles v ON o.vehicle_id = v.id
            JOIN clients c ON o.client_id = c.id
            {where}
            ORDER BY o.created_at DESC
        ''', params)
        results = [dict(row) for row in cursor.fetchall()]
        self.close()
        return results

    def update_order_status(self, order_id, new_status, note=None):
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('SELECT status FROM orders WHERE id = ?', (order_id,))
        row = cursor.fetchone()
        old_status = row['status'] if row else None
        cursor.execute('''
            UPDATE orders
            SET status = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (new_status, order_id))
        cursor.execute('''
            INSERT INTO order_status_history (order_id, old_status, new_status, note)
            VALUES (?, ?, ?, ?)
        ''', (order_id, old_status, new_status, note))
        conn.commit()
        self.close()

    def _reconcile_order_payments(self, order_id, cursor):
        self._ensure_order_payments_columns(cursor)
        cursor.execute('SELECT total_amount, client_id FROM orders WHERE id=?', (order_id,))
        order = cursor.fetchone()
        if not order:
            return
        total = float(order['total_amount'] or 0)
        client_id = order['client_id']
        cursor.execute('''
            SELECT id, amount, COALESCE(applied_amount, 0) as applied_amount, COALESCE(credit_amount, 0) as credit_amount
            FROM order_payments
            WHERE order_id=?
            ORDER BY paid_at ASC, id ASC
        ''', (order_id,))
        payments = [dict(row) for row in cursor.fetchall()]
        old_credit_total = sum(float(p.get('credit_amount') or 0) for p in payments)
        remaining = total
        new_paid_amount = 0.0
        new_credit_total = 0.0
        for payment in payments:
            amount = float(payment['amount'] or 0)
            applied = min(amount, max(remaining, 0.0))
            credit = amount - applied
            remaining = max(remaining - applied, 0.0)
            new_paid_amount += applied
            new_credit_total += credit
            cursor.execute('''
                UPDATE order_payments
                SET applied_amount=?, credit_amount=?
                WHERE id=?
            ''', (applied, credit, payment['id']))
        if new_paid_amount >= total and total > 0:
            status = PaymentStatus.PAID.value
        elif new_paid_amount > 0:
            status = PaymentStatus.PARTIAL.value
        else:
            status = PaymentStatus.UNPAID.value
        cursor.execute('''
            UPDATE orders
            SET paid_amount=?, payment_status=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        ''', (new_paid_amount, status, order_id))
        credit_delta = new_credit_total - old_credit_total
        if abs(credit_delta) > 0.000001:
            cursor.execute('SELECT COALESCE(balance, 0) FROM clients WHERE id=?', (client_id,))
            row = cursor.fetchone()
            if row:
                balance = float(row[0] or 0)
                new_balance = balance + credit_delta
                if new_balance > 1000000 or new_balance < -1000000:
                    self._recalculate_client_balance_from_scratch(client_id, cursor)
                else:
                    cursor.execute('UPDATE clients SET balance=? WHERE id=?', (new_balance, client_id))

    def _recalculate_client_balance_from_scratch(self, client_id, cursor):
        cursor.execute('''
            SELECT COALESCE(SUM(COALESCE(credit_amount, 0)), 0) as total_credit
            FROM order_payments
            WHERE order_id IN (SELECT id FROM orders WHERE client_id = ?)
        ''', (client_id,))
        row = cursor.fetchone()
        total_credit = float(row[0] or 0) if row else 0.0
        cursor.execute('UPDATE clients SET balance=? WHERE id=?', (total_credit, client_id))

    def update_order_payment(self, order_id, payment_amount, note=None):
        payment_amount = max(float(payment_amount or 0), 0.0)
        if payment_amount <= 0:
            return
        if payment_amount > 10000000:
            logger.warning("update_order_payment: payment amount %s exceeds max limit", payment_amount)
            return
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM orders WHERE id=?', (order_id,))
        if not cursor.fetchone():
            self.close()
            return
        cursor.execute('''
            INSERT INTO order_payments (order_id, amount, note)
            VALUES (?, ?, ?)
        ''', (order_id, payment_amount, note))
        self._reconcile_order_payments(order_id, cursor)
        conn.commit()
        self.close()

    def edit_order_payment(self, payment_id, amount, note=None):
        amount = max(float(amount or 0), 0.0)
        if amount > 10000000:
            return
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('SELECT order_id FROM order_payments WHERE id=?', (payment_id,))
        row = cursor.fetchone()
        if not row:
            self.close()
            return
        order_id = row['order_id']
        cursor.execute('UPDATE order_payments SET amount=?, note=? WHERE id=?', (amount, note, payment_id))
        self._reconcile_order_payments(order_id, cursor)
        conn.commit()
        self.close()

    def delete_order_payment(self, payment_id):
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('SELECT order_id FROM order_payments WHERE id=?', (payment_id,))
        row = cursor.fetchone()
        if not row:
            self.close()
            return
        order_id = row['order_id']
        cursor.execute('DELETE FROM order_payments WHERE id=?', (payment_id,))
        self._reconcile_order_payments(order_id, cursor)
        conn.commit()
        self.close()

    def update_client_balance(self, client_id, new_balance):
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('UPDATE clients SET balance=? WHERE id=?', (float(new_balance or 0), client_id))
        conn.commit()
        self.close()

    def _ensure_client_balance_column(self, cursor):
        self._add_column_if_missing(cursor, 'clients', 'balance', 'REAL DEFAULT 0')

    def get_order_details(self, order_id):
        conn = self.connect()
        cursor = conn.cursor()
        try:
            cursor.execute('''
                SELECT o.*, v.vin, v.brand, v.model, v.license_plate,
                       c.full_name, c.phone, c.balance AS client_balance
                FROM orders o
                JOIN vehicles v ON o.vehicle_id = v.id
                JOIN clients c ON o.client_id = c.id
                WHERE o.id = ?
            ''', (order_id,))
        except sqlite3.OperationalError as e:
            if 'no such column: c.balance' in str(e):
                self._ensure_client_balance_column(cursor)
                cursor.execute('''
                    SELECT o.*, v.vin, v.brand, v.model, v.license_plate,
                           c.full_name, c.phone, c.balance AS client_balance
                    FROM orders o
                    JOIN vehicles v ON o.vehicle_id = v.id
                    JOIN clients c ON o.client_id = c.id
                    WHERE o.id = ?
                ''', (order_id,))
            else:
                self.close()
                raise
        row = cursor.fetchone()
        if not row:
            self.close()
            return None
        order = dict(row)
        self.close()
        order['labor'] = self.get_order_labor(order_id)
        order['parts'] = self.get_order_parts(order_id)
        order['services'] = self.get_order_services(order_id)
        order['payments'] = self.get_order_payments(order_id)
        order['status_history'] = self.get_order_status_history(order_id)
        return order

    def get_order_status_history(self, order_id):
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT * FROM order_status_history
            WHERE order_id=?
            ORDER BY changed_at DESC, id DESC
        ''', (order_id,))
        rows = [dict(row) for row in cursor.fetchall()]
        self.close()
        return rows

    def get_order_payments(self, order_id):
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT * FROM order_payments
            WHERE order_id=?
            ORDER BY paid_at ASC, id ASC
        ''', (order_id,))
        rows = [dict(row) for row in cursor.fetchall()]
        self.close()
        return rows

    def export_order_to_excel(self, order_id, filename):
        import openpyxl
        order = self.get_order_details(order_id)
        if not order:
            raise ValueError('Заказ не найден')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Заказ {order['order_number']}"
        ws.append([f"Заказ-наряд №{order['order_number']}"])
        ws.append(["Клиент", order.get('full_name', '')])
        ws.append(["Телефон", order.get('phone', '')])
        ws.append(["Авто", f"{order.get('brand', '')} {order.get('model', '')}"])
        ws.append(["VIN", order.get('vin', '')])
        ws.append(["Номер", order.get('license_plate', '')])
        ws.append(["Статус", order.get('status', '')])
        ws.append(["Жалоба", order.get('complaint', '')])
        ws.append([])
        ws.append(["Запчасти"])
        ws.append(["Номер", "Описание", "Кол-во", "Цена", "Сумма", "Поставщик", "Статус"])
        for part in order['parts']:
            ws.append([part.get('part_number', ''), part.get('part_name', ''),
                       part.get('quantity', 0), part.get('selling_price', 0),
                       part.get('quantity', 0) * part.get('selling_price', 0),
                       part.get('supplier', ''), part.get('status', '')])
        ws.append([])
        ws.append(["Услуги"])
        ws.append(["Описание", "Кол-во", "Цена с НДС", "Сумма с НДС", "Комментарий"])
        for service in order['services']:
            ws.append([service.get('description', ''), service.get('quantity', 0),
                       service.get('price_with_vat', 0), service.get('subtotal_with_vat', 0),
                       service.get('notes', '')])
        ws.append([])
        ws.append(["Итого", order.get('total_amount', 0)])
        wb.save(filename)

    def export_order_to_pdf(self, order_id, filename):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        order = self.get_order_details(order_id)
        if not order:
            raise ValueError('Заказ не найден')
        font_name = 'Helvetica'
        try:
            pdfmetrics.registerFont(TTFont('Arial', r'C:\Windows\Fonts\arial.ttf'))
            font_name = 'Arial'
        except Exception:
            pass
        styles = getSampleStyleSheet()
        for style in styles.byName.values():
            style.fontName = font_name
        doc = SimpleDocTemplate(filename, pagesize=A4)
        story = [
            Paragraph(f"Заказ-наряд №{order['order_number']}", styles['Title']),
            Paragraph(f"Клиент: {order.get('full_name', '')}", styles['Normal']),
            Paragraph(f"Телефон: {order.get('phone', '')}", styles['Normal']),
            Paragraph(f"Авто: {order.get('brand', '')} {order.get('model', '')}; VIN: {order.get('vin', '')}; Номер: {order.get('license_plate', '')}", styles['Normal']),
            Paragraph(f"Жалоба: {order.get('complaint', '')}", styles['Normal']),
            Spacer(1, 12),
            Paragraph("Запчасти", styles['Heading2']),
        ]
        parts_data = [["Номер", "Описание", "Кол-во", "Цена", "Сумма"]]
        for part in order['parts']:
            parts_data.append([part.get('part_number', ''), part.get('part_name', ''),
                               part.get('quantity', 0), f"{part.get('selling_price', 0):.2f}",
                               f"{part.get('quantity', 0) * part.get('selling_price', 0):.2f}"])
        story.append(self._pdf_table(parts_data, font_name))
        story.extend([Spacer(1, 12), Paragraph("Услуги", styles['Heading2'])])
        services_data = [["Описание", "Кол-во", "Цена с НДС", "Сумма"]]
        for service in order['services']:
            services_data.append([service.get('description', ''), service.get('quantity', 0),
                                  f"{service.get('price_with_vat', 0):.2f}", f"{service.get('subtotal_with_vat', 0):.2f}"])
        story.append(self._pdf_table(services_data, font_name))
        story.extend([Spacer(1, 12), Paragraph(f"Итого: {order.get('total_amount', 0):.2f}", styles['Heading2'])])
        doc.build(story)

    def _pdf_table(self, data, font_name):
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        return table

    def export_order_to_html(self, order_id, filename):
        order = self.get_order_details(order_id)
        if not order:
            raise ValueError('Заказ не найден')
        parts_rows = ''.join(
            f"<tr><td>{escape(str(p.get('part_number') or ''))}</td>"
            f"<td>{escape(str(p.get('part_name') or ''))}</td>"
            f"<td>{p.get('quantity', 0)}</td><td>{p.get('selling_price', 0):.2f}</td>"
            f"<td>{p.get('quantity', 0) * p.get('selling_price', 0):.2f}</td></tr>"
            for p in order['parts']
        )
        service_rows = ''.join(
            f"<tr><td>{escape(str(s.get('description') or ''))}</td>"
            f"<td>{s.get('quantity', 0)}</td><td>{s.get('price_with_vat', 0):.2f}</td>"
            f"<td>{s.get('subtotal_with_vat', 0):.2f}</td></tr>"
            for s in order['services']
        )
        html = f"""<!doctype html>
<html lang="ru">
<head><meta charset="utf-8"><title>Заказ-наряд №{order['order_number']}</title>
<style>body{{font-family:Arial,sans-serif;margin:32px;color:#17202a}}h1{{font-size:24px}}table{{border-collapse:collapse;width:100%;margin:16px 0}}th,td{{border:1px solid #ccd1d1;padding:8px;text-align:left}}th{{background:#ecf0f1}}.total{{text-align:right;font-size:18px;font-weight:bold}}@media print{{button{{display:none}}}}</style>
</head>
<body>
<button onclick="window.print()">Печать</button>
<h1>Заказ-наряд №{order['order_number']}</h1>
<p><b>Клиент:</b> {escape(str(order.get('full_name') or ''))} | <b>Телефон:</b> {escape(str(order.get('phone') or ''))}</p>
<p><b>Авто:</b> {escape(str(order.get('brand') or ''))} {escape(str(order.get('model') or ''))} | <b>VIN:</b> {escape(str(order.get('vin') or ''))} | <b>Номер:</b> {escape(str(order.get('license_plate') or ''))}</p>
<p><b>Жалоба:</b> {escape(str(order.get('complaint') or ''))}</p>
<h2>Запчасти</h2>
<table><tr><th>Номер</th><th>Описание</th><th>Кол-во</th><th>Цена</th><th>Сумма</th></tr>{parts_rows}</table>
<h2>Услуги</h2>
<table><tr><th>Описание</th><th>Кол-во</th><th>Цена с НДС</th><th>Сумма</th></tr>{service_rows}</table>
<p class="total">Итого: {order.get('total_amount', 0):.2f}</p>
</body></html>"""
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(html)

    # ========== ФИНАНСЫ И АНАЛИТИКА ==========
    def get_finance_summary(self, period='day'):
        date_from = self._period_start(period)
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT COUNT(*) AS order_count,
                   COALESCE(SUM(o.total_amount), 0) AS revenue,
                   COALESCE(SUM(o.paid_amount), 0) AS paid,
                   COALESCE(SUM(o.total_amount - COALESCE(o.paid_amount, 0)), 0) AS debt,
                   COALESCE(SUM(parts_cost.cost), 0) AS parts_cost
            FROM orders o
            LEFT JOIN (
                SELECT order_id, SUM(quantity * purchase_price) AS cost
                FROM order_parts GROUP BY order_id
            ) parts_cost ON parts_cost.order_id = o.id
            WHERE o.created_at >= ?
        ''', (date_from,))
        summary = dict(cursor.fetchone())
        summary['profit'] = float(summary['revenue'] or 0) - float(summary['parts_cost'] or 0)
        summary['period'] = period
        summary['date_from'] = date_from
        self.close()
        return summary

    def get_order_profit_rows(self, period='month'):
        date_from = self._period_start(period)
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT o.id, o.order_number, o.created_at, c.full_name,
                   v.brand, v.model, v.license_plate,
                   o.total_amount, COALESCE(o.paid_amount, 0) AS paid_amount,
                   o.payment_status,
                   COALESCE(parts_cost.cost, 0) AS parts_cost,
                   o.total_amount - COALESCE(parts_cost.cost, 0) AS profit,
                   o.total_amount - COALESCE(o.paid_amount, 0) AS debt
            FROM orders o
            JOIN clients c ON o.client_id = c.id
            JOIN vehicles v ON o.vehicle_id = v.id
            LEFT JOIN (
                SELECT order_id, SUM(quantity * purchase_price) AS cost
                FROM order_parts GROUP BY order_id
            ) parts_cost ON parts_cost.order_id = o.id
            WHERE o.created_at >= ?
            ORDER BY o.created_at DESC
        ''', (date_from,))
        rows = [dict(row) for row in cursor.fetchall()]
        self.close()
        return rows

    def get_popular_services(self, period='month', limit=10):
        date_from = self._period_start(period)
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT s.description, SUM(s.quantity) AS quantity,
                   COUNT(DISTINCT s.order_id) AS order_count,
                   SUM(s.subtotal_with_vat) AS revenue
            FROM order_services s
            JOIN orders o ON o.id = s.order_id
            WHERE o.created_at >= ?
            GROUP BY s.description
            ORDER BY quantity DESC, revenue DESC LIMIT ?
        ''', (date_from, limit))
        rows = [dict(row) for row in cursor.fetchall()]
        self.close()
        return rows

    def get_popular_parts(self, period='month', limit=10):
        date_from = self._period_start(period)
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT p.part_name, p.part_number,
                   SUM(p.quantity) AS quantity, COUNT(DISTINCT p.order_id) AS order_count,
                   SUM(p.quantity * p.selling_price) AS revenue,
                   SUM(p.quantity * p.purchase_price) AS cost,
                   SUM(p.quantity * (p.selling_price - p.purchase_price)) AS profit
            FROM order_parts p
            JOIN orders o ON o.id = p.order_id
            WHERE o.created_at >= ?
            GROUP BY p.part_name, p.part_number
            ORDER BY quantity DESC, revenue DESC LIMIT ?
        ''', (date_from, limit))
        rows = [dict(row) for row in cursor.fetchall()]
        self.close()
        return rows

    def get_debts(self):
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT o.id, o.order_number, o.created_at, c.full_name, c.phone,
                   v.brand, v.model, v.license_plate,
                   o.total_amount, COALESCE(o.paid_amount, 0) AS paid_amount,
                   o.total_amount - COALESCE(o.paid_amount, 0) AS debt,
                   o.payment_status
            FROM orders o
            JOIN clients c ON o.client_id = c.id
            JOIN vehicles v ON o.vehicle_id = v.id
            WHERE o.total_amount > COALESCE(o.paid_amount, 0)
            ORDER BY debt DESC, o.created_at DESC
        ''')
        rows = [dict(row) for row in cursor.fetchall()]
        self.close()
        return rows

    def export_finance_report_to_excel(self, filename, period='month'):
        import openpyxl
        summary = self.get_finance_summary(period)
        orders = self.get_order_profit_rows(period)
        services = self.get_popular_services(period, 25)
        parts = self.get_popular_parts(period, 25)
        debts = self.get_debts()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Сводка'
        for row in [['Период', period], ['Дата с', summary['date_from']],
                    ['Заказов', summary['order_count']], ['Выручка', summary['revenue']],
                    ['Оплачено', summary['paid']], ['Долг', summary['debt']],
                    ['Себестоимость запчастей', summary['parts_cost']], ['Прибыль', summary['profit']]]:
            ws.append(row)
        self._append_sheet(wb, 'Заказы', orders)
        self._append_sheet(wb, 'Услуги', services)
        self._append_sheet(wb, 'Запчасти', parts)
        self._append_sheet(wb, 'Долги', debts)
        wb.save(filename)

    def _append_sheet(self, wb, title, rows):
        ws = wb.create_sheet(title)
        if not rows:
            ws.append(['Нет данных'])
            return
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h) for h in headers])

    def _period_start(self, period):
        today = datetime.now()
        if period == 'day':
            start = today.replace(hour=0, minute=0, second=0, microsecond=0)
        elif period == 'week':
            start = (today - timedelta(days=today.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        else:
            start = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        return start.strftime('%Y-%m-%d %H:%M:%S')

    # ========== РАБОТЫ (старые) ==========
    def add_labor(self, order_id, description, price, quantity=1):
        conn = self.connect()
        cursor = conn.cursor()
        subtotal = price * quantity
        cursor.execute('''
            INSERT INTO order_labor (order_id, description, price, quantity, subtotal)
            VALUES (?, ?, ?, ?, ?)
        ''', (order_id, description, price, quantity, subtotal))
        conn.commit()
        labor_id = cursor.lastrowid
        self._update_order_total(conn, order_id)
        self.close()
        return labor_id

    def get_order_labor(self, order_id):
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM order_labor WHERE order_id = ?', (order_id,))
        rows = [dict(r) for r in cursor.fetchall()]
        self.close()
        return rows

    # ========== ЗАПЧАСТИ ==========
    def add_part(self, order_id, part_name, quantity, selling_price,
                 purchase_price=0.0, part_number=None, supplier=None,
                 status=PartStatus.NEEDS_ORDER.value, notes=None):
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO order_parts
            (order_id, part_name, part_number, quantity, purchase_price,
             selling_price, status, supplier, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (order_id, part_name, part_number, quantity, purchase_price,
              selling_price, status, supplier, notes))
        conn.commit()
        part_id = cursor.lastrowid
        self._update_order_total(conn, order_id)
        self.close()
        return part_id

    def delete_part(self, part_id):
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('SELECT order_id FROM order_parts WHERE id=?', (part_id,))
        row = cursor.fetchone()
        if not row:
            self.close()
            return
        order_id = row['order_id']
        cursor.execute('DELETE FROM order_parts WHERE id=?', (part_id,))
        conn.commit()
        self._update_order_total(conn, order_id)
        self.close()

    def update_part(self, part_id, part_name=None, part_number=None, quantity=None,
                    selling_price=None, purchase_price=None, supplier=None, status=None, notes=None):
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM order_parts WHERE id=?', (part_id,))
        current = cursor.fetchone()
        if not current:
            self.close()
            return
        new = {
            'part_name': part_name if part_name is not None else current['part_name'],
            'part_number': part_number if part_number is not None else current['part_number'],
            'quantity': quantity if quantity is not None else current['quantity'],
            'selling_price': selling_price if selling_price is not None else current['selling_price'],
            'purchase_price': purchase_price if purchase_price is not None else current['purchase_price'],
            'supplier': supplier if supplier is not None else current['supplier'],
            'status': status if status is not None else current['status'],
            'notes': notes if notes is not None else current['notes'],
        }
        cursor.execute('''
            UPDATE order_parts
            SET part_name=?, part_number=?, quantity=?, selling_price=?,
                purchase_price=?, supplier=?, status=?, notes=?
            WHERE id=?
        ''', (new['part_name'], new['part_number'], new['quantity'],
              new['selling_price'], new['purchase_price'],
              new['supplier'], new['status'], new['notes'], part_id))
        conn.commit()
        self._update_order_total(conn, current['order_id'])
        self.close()

    def get_order_parts(self, order_id):
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM order_parts WHERE order_id = ?', (order_id,))
        rows = [dict(r) for r in cursor.fetchall()]
        self.close()
        return rows

    # ========== УСЛУГИ ==========
    def add_service(self, order_id, description, price_with_vat, quantity=1):
        """Добавить услугу. price_with_vat — конечная сумма с НДС за единицу."""
        price_without_vat, vat_amount, _ = split_unit_price_with_vat(price_with_vat)
        subtotal_without_vat, vat_total, subtotal_with_vat = line_totals_with_vat(price_with_vat, quantity)
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO order_services
            (order_id, description, price_with_vat, vat_amount, price_without_vat,
             quantity, subtotal_with_vat, subtotal_without_vat, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (order_id, description, price_with_vat, vat_amount, price_without_vat,
              quantity, subtotal_with_vat, subtotal_without_vat, ''))
        conn.commit()
        service_id = cursor.lastrowid
        self._update_order_total(conn, order_id)
        self.close()
        return service_id

    def delete_service(self, service_id):
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('SELECT order_id FROM order_services WHERE id=?', (service_id,))
        row = cursor.fetchone()
        if not row:
            self.close()
            return
        order_id = row['order_id']
        cursor.execute('DELETE FROM order_services WHERE id=?', (service_id,))
        conn.commit()
        self._update_order_total(conn, order_id)
        self.close()

    def update_service(self, service_id, description=None, price_with_vat=None, quantity=None, notes=None):
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM order_services WHERE id=?', (service_id,))
        current = cursor.fetchone()
        if not current:
            self.close()
            return
        new_desc = description if description is not None else current['description']
        new_qty = quantity if quantity is not None else current['quantity']
        new_price = price_with_vat if price_with_vat is not None else current['price_with_vat']
        new_notes = notes if notes is not None else current['notes']
        price_without_vat, vat_amount, _ = split_unit_price_with_vat(new_price)
        subtotal_without_vat, vat_total, subtotal_with_vat = line_totals_with_vat(new_price, new_qty)
        cursor.execute('''
            UPDATE order_services
            SET description=?, price_with_vat=?, vat_amount=?, price_without_vat=?,
                quantity=?, subtotal_with_vat=?, subtotal_without_vat=?, notes=?
            WHERE id=?
        ''', (new_desc, new_price, vat_amount, price_without_vat, new_qty,
              subtotal_with_vat, subtotal_without_vat, new_notes, service_id))
        conn.commit()
        self._update_order_total(conn, current['order_id'])
        self.close()

    def get_order_services(self, order_id):
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM order_services WHERE order_id = ?', (order_id,))
        rows = [dict(r) for r in cursor.fetchall()]
        self.close()
        return rows

    def update_part_status(self, part_id, new_status, purchase_price=None, received_at=None):
        conn = self.connect()
        cursor = conn.cursor()
        if received_at:
            cursor.execute('''
                UPDATE order_parts
                SET status=?, purchase_price=COALESCE(?, purchase_price), received_at=?
                WHERE id=?
            ''', (new_status, purchase_price, received_at, part_id))
        else:
            cursor.execute('''
                UPDATE order_parts
                SET status=?, purchase_price=COALESCE(?, purchase_price)
                WHERE id=?
            ''', (new_status, purchase_price, part_id))
        conn.commit()
        self.close()

    def save_order_items(self, order_id, parts, services, deleted_ids=None):
        return self.orders.save_order_items(order_id, parts, services, deleted_ids)

    def _update_order_total(self, conn, order_id, commit=True):
        cursor = conn.cursor()
        cursor.execute('SELECT COALESCE(SUM(subtotal),0) FROM order_labor WHERE order_id=?', (order_id,))
        total_labor = cursor.fetchone()[0]
        cursor.execute('SELECT COALESCE(SUM(selling_price * quantity),0) FROM order_parts WHERE order_id=?', (order_id,))
        total_parts = cursor.fetchone()[0]
        cursor.execute('SELECT COALESCE(SUM(selling_price * quantity * ?),0) FROM order_parts WHERE order_id=?', (VAT_RATE, order_id))
        total_parts_vat = cursor.fetchone()[0]
        cursor.execute('SELECT COALESCE(SUM(subtotal_with_vat),0) FROM order_services WHERE order_id=?', (order_id,))
        total_services = cursor.fetchone()[0]
        total = round(total_labor + total_parts + total_parts_vat + total_services, 2)
        cursor.execute('''
            UPDATE orders SET total_labor=?, total_parts=?, total_services=?, total_amount=?,
                updated_at=CURRENT_TIMESTAMP WHERE id=?
        ''', (total_labor, total_parts, total_services, total, order_id))
        self._reconcile_order_payments(order_id, cursor)
        if commit:
            conn.commit()

    # ========== ПОСТАВЩИКИ ==========
    def add_supplier(self, name):
        conn = self.connect()
        cursor = conn.cursor()
        try:
            cursor.execute('INSERT INTO suppliers (name) VALUES (?)', (name,))
            conn.commit()
            supplier_id = cursor.lastrowid
            self.close()
            return supplier_id
        except sqlite3.IntegrityError:
            self.close()
            return None

    def get_all_suppliers(self):
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM suppliers ORDER BY name')
        suppliers = [dict(row) for row in cursor.fetchall()]
        self.close()
        return suppliers

    def delete_supplier(self, supplier_id):
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM suppliers WHERE id=?', (supplier_id,))
        conn.commit()
        self.close()
